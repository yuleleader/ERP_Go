# -*- coding: utf-8 -*-
"""订单批量导入：Excel → 临时表(order_imports) → 勾选审核 → 合并正式订单表。

流程设计（与用户确认）：
1. 系统运维新增「数据导入」入口；模板含列名 + 案例数据（示例行以"示例"开头，导入时自动跳过）。
2. 上传 Excel 后先写入临时表（同批同 batch_no），不直接进正式表。
3. 每行实时计算异常规则（errors JSON 数组），异常订单保留在临时表并逐行展示异常原因。
4. 支持勾选/全选后「审核合并」：仅无异常行可合并；合并时自动生成符合业务规则的追溯码（order_id），
   写入正式 orders 表后临时记录移除。
5. 异常订单可点击编辑，保存后重新计算异常，仍需再次审核才能合并。
6. 可见范围：销售只看自己导入的，老板看全部；可单条/批量删除。

异常规则全集：
- 平台订单号为空 / 库内已存在 / 文件内重复
- 销售金额、运费：非数字或负数
- 下单时间：格式错误、晚于当前日期
- 发货时间：格式错误、晚于当前日期
- 下单时间晚于发货时间（两者都有值时校验）
- 发货状态非法、生产状态非法
- 网店不存在 / 匹配到多个 / 已关店
- 已发货/虚拟发货但生产未完成
- 已退货/退款但退款备注为空
- 已发货但物流单号为空
"""
import json
import random
import string
from io import BytesIO
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from urllib.parse import quote

from ..core.database import get_db
from ..core.security import get_current_active_user, ensure_data_permission
from ..models.models import Order, Shop, User, Product, OrderImport, OperationLog, beijing_now
from ..api.orders import generate_order_id, calc_order_gross_profit, calculate_order_days
from ..services.notification_service import NotificationService

router = APIRouter(prefix="/api/order-imports", tags=["订单导入"])

# ==================== 常量 ====================
# 模板列（中文表头）：列名 + 案例数据一起下载
TEMPLATE_HEADERS = [
    "平台订单号", "网店名称", "网店账号", "商品名称", "销售金额",
    "运费", "发货状态", "生产状态", "快递公司", "运单号",
    "收货地址", "备注", "退款备注", "下单时间", "发货时间",
]
# 发货状态：中文 → 英文枚举
SHIPPING_STATUS_MAP = {
    "待发货": "pending",
    "已发货": "shipped",
    "虚拟发货": "virtual",
    "已退货/退款": "refunded",
    "已退货退款": "refunded",
    "已退货": "refunded",
    "已退款": "refunded",
}
SHIPPING_STATUS_CN = {v: k for k, v in SHIPPING_STATUS_MAP.items()}
# 生产状态：中文 → 英文枚举
PRODUCE_STATUS_MAP = {
    "未生产": "unproduce",
    "生产中": "producing",
    "生产完成": "produced",
}
PRODUCE_STATUS_CN = {v: k for k, v in PRODUCE_STATUS_MAP.items()}

# 模板示例行（平台订单号以"示例"开头，导入时自动跳过）
SAMPLE_ROWS = [
    [
        "示例-20260810-001", "示例店铺", "shop@example.com", "示例商品A", 99.9,
        5, "待发货", "未生产", "顺丰", "SF1234567890",
        "广东省深圳市南山区科技园路1号", "首次导入示例数据", "", "2026-08-10", "",
    ],
    [
        "示例-20260810-002", "示例店铺", "shop@example.com", "示例商品B", 50,
        3, "已发货", "未生产", "中通", "ZT9876543210",
        "广东省广州市天河区路2号", "异常示例：已发货但生产未完成", "", "2026-08-09", "2026-08-10",
    ],
]

ALLOWED_EXTENSIONS = ("xlsx", "xls")
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_ROWS = 5000  # 单次导入最大行数


# ==================== Schema ====================
class MergeRequest(BaseModel):
    ids: List[int]


class DeleteBatchRequest(BaseModel):
    ids: List[int]


# ==================== 工具函数 ====================
def _parse_excel_rows(content: bytes, ext: str) -> list:
    """解析 Excel 为字符串二维数组（含表头行）。支持 .xlsx（openpyxl）与 .xls（xlrd）。"""
    if ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([
                "" if v is None
                else (v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, datetime) else str(v).strip())
                for v in row
            ])
    else:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        ws = book.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                if ws.cell_type(r, c) == 3 and isinstance(v, float):
                    v = xlrd.xldate_as_datetime(v, book.datemode)
                row.append("" if v is None else str(v).strip())
            rows.append(row)
    return rows


def _parse_dt(s: str):
    """尝试解析日期字符串为 datetime；失败返回 None。"""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _norm_dt(s: str) -> str:
    """规范化日期字符串（用于入库）；无法解析时返回原值。"""
    dt = _parse_dt(s)
    if dt is None:
        return (s or "").strip()
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        return dt.strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _money_ok(s) -> bool:
    s = (s or "").strip()
    if not s:
        return True  # 空值不算错（可编辑补填）
    try:
        return float(s) >= 0
    except ValueError:
        return False


def _gen_batch_no() -> str:
    return "IMP" + beijing_now().strftime("%Y%m%d%H%M%S") + "".join(random.choices(string.digits, k=4))


def _row_to_dict(row: OrderImport, user_map=None) -> dict:
    """临时表行 → 前端字典（errors 解析为数组）。"""
    try:
        errors = json.loads(row.errors) if row.errors else []
    except (TypeError, ValueError):
        errors = []
    if not isinstance(errors, list):
        errors = []
    return {
        "id": row.id,
        "batch_no": row.batch_no,
        "platform_order_no": row.platform_order_no or "",
        "shop_name": row.shop_name or "",
        "shop_account": row.shop_account or "",
        "shop_id": row.shop_id or "",
        "product_name": row.product_name or "",
        "sales_amount": row.sales_amount or "",
        "freight": row.freight or "",
        "shipping_status": row.shipping_status or "pending",
        "shipping_status_text": SHIPPING_STATUS_CN.get(row.shipping_status, row.shipping_status or ""),
        "produce_status": row.produce_status or "unproduce",
        "produce_status_text": PRODUCE_STATUS_CN.get(row.produce_status, row.produce_status or ""),
        "logistics_company": row.logistics_company or "",
        "logistics_no": row.logistics_no or "",
        "receiver_address": row.receiver_address or "",
        "remark": row.remark or "",
        "refund_note": row.refund_note or "",
        "order_time": row.order_time or "",
        "shipping_time": row.shipping_time or "",
        "errors": errors,
        "imported_by": row.imported_by or "",
        "imported_by_name": (user_map or {}).get(row.imported_by or "", row.imported_by or ""),
        "import_time": row.import_time.strftime("%Y-%m-%d %H:%M:%S") if row.import_time else None,
        "status": row.status or "pending",
        "merged_order_id": row.merged_order_id or "",
    }


async def compute_import_errors(db: AsyncSession, row: OrderImport, batch_seen: Optional[set] = None) -> list:
    """计算单行临时订单的全部异常（返回异常字符串列表）。batch_seen 非空时追加"文件内重复"检测（导入时用）。"""
    errors = []
    pno = (row.platform_order_no or "").strip()

    # 1) 平台订单号
    if not pno:
        errors.append("平台订单号不能为空")
    else:
        exists = (await db.execute(
            select(Order).where(Order.platform_order_no == pno)
        )).scalar_one_or_none()
        if exists:
            errors.append("平台订单号已存在于正式订单")
        if batch_seen is not None and pno in batch_seen:
            errors.append("文件内平台订单号重复")

    # 2) 金额
    for label, val in (("销售金额", row.sales_amount), ("运费", row.freight)):
        s = (val or "").strip()
        if s and not _money_ok(s):
            errors.append(f"{label}格式不正确或为负数")

    # 3) 时间：下单时间
    ot_dt = _parse_dt(row.order_time)
    if (row.order_time or "").strip() and ot_dt is None:
        errors.append("下单时间格式不正确")
    else:
        if ot_dt is not None and ot_dt.date() > beijing_now().date():
            errors.append("下单时间晚于当前日期")

    # 4) 时间：发货时间
    st_dt = _parse_dt(row.shipping_time)
    if (row.shipping_time or "").strip() and st_dt is None:
        errors.append("发货时间格式不正确")
    else:
        if st_dt is not None and st_dt.date() > beijing_now().date():
            errors.append("发货时间晚于当前日期")

    # 5) 时间顺序：下单晚于发货
    if ot_dt is not None and st_dt is not None and ot_dt > st_dt:
        errors.append("下单时间晚于发货时间")

    # 6) 状态枚举
    if row.shipping_status not in ("pending", "shipped", "virtual", "refunded"):
        errors.append("发货状态不合法")
    if row.produce_status not in ("unproduce", "producing", "produced"):
        errors.append("生产状态不合法")

    # 7) 网店：定位正式 shop_id
    shop_id, shop_status = await _resolve_shop_id(db, row)
    row.shop_id = shop_id
    if shop_status == "multiple":
        errors.append("匹配到多个网店，请同时填写名称与账号")
    elif shop_status == "none":
        if not (row.shop_name or "").strip() and not (row.shop_account or "").strip():
            errors.append("网店名称/网店账号未填写")
        else:
            errors.append("网店不存在，请检查名称/账号")
    else:  # ok：唯一匹配成功，补充「已关店」校验（与手工建单一致）
        shop = await db.get(Shop, shop_id)
        if shop and shop.status == "closed":
            errors.append("该网店已关店，无法导入")

    # 8) 业务规则
    if row.shipping_status in ("shipped", "virtual") and row.produce_status != "produced":
        errors.append("已发货/虚拟发货但生产未完成")
    if row.shipping_status == "refunded" and not (row.refund_note or "").strip():
        errors.append("已退货/退款但退款备注为空")
    if row.shipping_status == "shipped" and not (row.logistics_no or "").strip():
        errors.append("已发货但物流单号为空")

    return errors


async def _resolve_shop_id(db: AsyncSession, row: OrderImport):
    """按名称+账号组合优先、其次账号、再次名称定位正式网店 shop_id。

    返回 (shop_id, status)：
      - (shop_id, "ok")       唯一匹配成功
      - (None, "multiple")    账号或名称命中多个网店（需同时填名称+账号）
      - (None, "none")        未匹配到任何网店
    """
    shop_name = (row.shop_name or "").strip()
    shop_account = (row.shop_account or "").strip()
    if not shop_name and not shop_account:
        return (None, "none")

    if shop_name and shop_account:
        s = (await db.execute(
            select(Shop).where(Shop.shop_name == shop_name, Shop.shop_account == shop_account)
        )).scalar_one_or_none()
        if s:
            return (s.shop_id, "ok")
        # 组合未匹配：退化为账号/名称分别尝试（组合匹配不到的多半是账号或名称写错）
    if shop_account:
        rows = (await db.execute(select(Shop).where(Shop.shop_account == shop_account))).scalars().all()
        if len(rows) > 1:
            return (None, "multiple")
        if len(rows) == 1:
            return (rows[0].shop_id, "ok")
    if shop_name:
        rows = (await db.execute(select(Shop).where(Shop.shop_name == shop_name))).scalars().all()
        if len(rows) > 1:
            return (None, "multiple")
        if len(rows) == 1:
            return (rows[0].shop_id, "ok")
    return (None, "none")


def _check_import_perm(current_user):
    if current_user.role not in ("boss", "sales"):
        raise HTTPException(status_code=403, detail="您没有权限使用数据导入功能")


def _check_row_visible(db_row: OrderImport, current_user):
    """销售只能操作自己导入的行；老板可操作全部。"""
    if current_user.role != "boss" and db_row.imported_by != current_user.username:
        raise HTTPException(status_code=403, detail="只能操作自己导入的数据")


# ==================== 接口 ====================
@router.get("/template")
async def download_template(current_user=Depends(get_current_active_user)):
    """下载 Excel 导入模板（含列名 + 案例数据；示例行导入时自动跳过）。"""
    _check_import_perm(current_user)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导入模板"
    ws.append(TEMPLATE_HEADERS)
    for r in SAMPLE_ROWS:
        ws.append(r)
    widths = [18, 14, 22, 18, 12, 10, 12, 12, 14, 20, 34, 20, 14, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    # 表头加粗
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"订单导入模板_{beijing_now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/import")
async def import_orders(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """上传 Excel → 解析 → 写入临时表（同批 batch_no），每行计算异常。返回批次统计。"""
    _check_import_perm(current_user)
    ensure_data_permission(current_user, '/order-imports', 'add')

    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式的 Excel 文件")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="文件过大，请控制在 10MB 以内")

    try:
        rows = _parse_excel_rows(content, ext)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败：{e}")

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有可导入的数据行")

    # 表头校验
    header = [str(h or "").strip() for h in rows[0]]
    header_index = {}
    for idx, h in enumerate(header):
        header_index[h] = idx
    if "平台订单号" not in header_index:
        raise HTTPException(status_code=400, detail="模板表头缺少「平台订单号」列，请使用下载的模板")

    # 数据行（跳过空行 + 示例行）
    data_rows = []
    for r in rows[1:]:
        if not any((str(v or "").strip() for v in r)):
            continue
        platform_no = str(r[header_index.get("平台订单号", 0)] or "").strip()
        if platform_no.startswith("示例"):
            continue
        data_rows.append(r)
    if not data_rows:
        raise HTTPException(status_code=400, detail="文件中没有可导入的数据行（示例行已自动跳过）")
    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"单次最多导入 {MAX_ROWS} 行，当前 {len(data_rows)} 行")

    batch_no = _gen_batch_no()

    def col(r, name):
        i = header_index.get(name)
        return "" if i is None or i >= len(r) else str(r[i] or "").strip()

    imported = 0
    # 文件内重复：统计出现次数 >=2 的平台订单号集合（这些行全部标记"文件内重复"）
    from collections import Counter
    pno_counts = Counter(col(r, "平台订单号") for r in data_rows)
    dup_pnos = {p for p, c in pno_counts.items() if c >= 2}
    created_rows = []
    for r in data_rows:
        platform_no = col(r, "平台订单号")
        # 状态中文 → 英文
        status_cn = col(r, "发货状态")
        shipping_status = SHIPPING_STATUS_MAP.get(status_cn, "pending")
        produce_cn = col(r, "生产状态")
        produce_status = PRODUCE_STATUS_MAP.get(produce_cn, "unproduce")

        row = OrderImport(
            batch_no=batch_no,
            platform_order_no=platform_no,
            shop_name=col(r, "网店名称"),
            shop_account=col(r, "网店账号"),
            product_name=col(r, "商品名称"),
            sales_amount=col(r, "销售金额"),
            freight=col(r, "运费"),
            shipping_status=shipping_status,
            produce_status=produce_status,
            logistics_company=col(r, "快递公司"),
            logistics_no=col(r, "运单号"),
            receiver_address=col(r, "收货地址"),
            remark=col(r, "备注"),
            refund_note=col(r, "退款备注"),
            order_time=_norm_dt(col(r, "下单时间")),
            shipping_time=_norm_dt(col(r, "发货时间")),
            imported_by=current_user.username,
        )
        db.add(row)
        created_rows.append(row)
        imported += 1

    # 批量提交后逐行计算异常（需要 row.id；文件内重复用 dup_pnos 集合）
    await db.commit()
    for row in created_rows:
        errors = await compute_import_errors(db, row, batch_seen=dup_pnos)
        row.errors = json.dumps(errors, ensure_ascii=False)
    await db.commit()

    # 操作日志
    db.add(OperationLog(
        username=current_user.username,
        operation_type="订单批量导入",
        operation_content=f"导入临时表 {imported} 行（批次 {batch_no}）",
    ))
    await db.commit()

    return {"batch_no": batch_no, "total": imported, "message": f"成功导入 {imported} 行到临时表（批次 {batch_no}）"}


@router.get("/")
async def list_order_imports(
    batch_no: str = Query(None, description="批次号（模糊）"),
    keyword: str = Query(None, description="平台订单号/商品名称（模糊）"),
    only_abnormal: bool = Query(False, description="仅看有异常的行"),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """临时表列表：销售只看自己导入的，老板看全部。"""
    _check_import_perm(current_user)
    query = select(OrderImport)
    if current_user.role != "boss":
        query = query.where(OrderImport.imported_by == current_user.username)
    if batch_no:
        query = query.where(OrderImport.batch_no.like(f"%{batch_no.strip()}%"))
    if keyword:
        kw = keyword.strip()
        query = query.where(or_(
            OrderImport.platform_order_no.like(f"%{kw}%"),
            OrderImport.product_name.like(f"%{kw}%"),
        ))
    if only_abnormal:
        # 仅看有异常的行：errors 非空且不是空数组 "[]"
        query = query.where(and_(
            OrderImport.errors.isnot(None),
            OrderImport.errors != "[]",
        ))

    from sqlalchemy import func as sa_func
    count_q = select(sa_func.count()).select_from(query.subquery())
    total_count = (await db.execute(count_q)).scalar() or 0

    rows = (await db.execute(
        query.order_by(OrderImport.id.desc()).offset(skip).limit(limit)
    )).scalars().all()

    user_map = {}
    for u in (await db.execute(select(User))).scalars().all():
        user_map[u.username] = u.real_name or u.username

    return {
        "total": total_count,
        "items": [_row_to_dict(r, user_map) for r in rows],
    }


@router.put("/{row_id}")
async def update_order_import(
    row_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """编辑临时表行：保存后重新计算全部异常，仍留在临时表，需再次审核。"""
    _check_import_perm(current_user)
    ensure_data_permission(current_user, '/order-imports', 'edit')
    row = (await db.execute(select(OrderImport).where(OrderImport.id == row_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="临时订单不存在")
    _check_row_visible(row, current_user)

    editable = [
        "platform_order_no", "shop_name", "shop_account", "product_name", "sales_amount",
        "freight", "shipping_status", "produce_status", "logistics_company", "logistics_no",
        "receiver_address", "remark", "refund_note", "order_time", "shipping_time",
    ]
    for k in editable:
        if k in payload:
            v = payload[k]
            setattr(row, k, "" if v is None else str(v).strip())
    # 状态字段允许中文值 → 转英文枚举
    if "shipping_status" in payload:
        val = str(payload.get("shipping_status") or "").strip()
        row.shipping_status = SHIPPING_STATUS_MAP.get(val, val if val in ("pending", "shipped", "virtual", "refunded") else "pending")
    if "produce_status" in payload:
        val = str(payload.get("produce_status") or "").strip()
        row.produce_status = PRODUCE_STATUS_MAP.get(val, val if val in ("unproduce", "producing", "produced") else "unproduce")
    # 时间规范化
    if "order_time" in payload:
        row.order_time = _norm_dt(str(payload.get("order_time") or ""))
    if "shipping_time" in payload:
        row.shipping_time = _norm_dt(str(payload.get("shipping_time") or ""))

    errors = await compute_import_errors(db, row)
    row.errors = json.dumps(errors, ensure_ascii=False)
    await db.commit()

    db.add(OperationLog(
        username=current_user.username,
        operation_type="订单导入编辑",
        operation_content=f"编辑临时订单 #{row.id}（{row.platform_order_no or ''}）",
    ))
    await db.commit()

    return _row_to_dict(row)


@router.post("/merge")
async def merge_order_imports(
    payload: MergeRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """审核合并：仅无异常行可合并。逐行生成追溯码写入正式 orders 表，成功后临时记录删除。"""
    _check_import_perm(current_user)
    ensure_data_permission(current_user, '/order-imports', 'edit')
    ids = payload.ids
    if not ids:
        raise HTTPException(status_code=400, detail="请先勾选要合并的订单")

    rows = []
    for rid in ids:
        row = (await db.execute(select(OrderImport).where(OrderImport.id == rid))).scalar_one_or_none()
        if not row:
            continue
        _check_row_visible(row, current_user)
        rows.append(row)

    if not rows:
        raise HTTPException(status_code=400, detail="勾选的订单不存在或无权操作")

    # 合并前逐行重新校验（防编辑/并发后异常变化）
    merged_ok, merged_fail = [], []
    for row in rows:
        errors = await compute_import_errors(db, row)
        row.errors = json.dumps(errors, ensure_ascii=False)
        if errors:
            merged_fail.append({"id": row.id, "platform_order_no": row.platform_order_no, "errors": errors})
            continue
        # 网店必须已定位（compute 已写 shop_id）
        shop = (await db.execute(select(Shop).where(Shop.shop_id == row.shop_id))).scalar_one_or_none()
        if not shop:
            merged_fail.append({"id": row.id, "platform_order_no": row.platform_order_no, "errors": ["网店不存在"]})
            continue

        # 生成追溯码（order_id = 导入人账号+年月日+网店账号+平台订单号+6位随机数），碰撞重试
        order_id = None
        for _ in range(5):
            candidate = await generate_order_id(row.imported_by, shop.shop_account, row.platform_order_no)
            exists = (await db.execute(select(Order).where(Order.order_id == candidate))).scalar_one_or_none()
            if not exists:
                order_id = candidate
                break
        if not order_id:
            merged_fail.append({"id": row.id, "platform_order_no": row.platform_order_no, "errors": ["追溯码生成冲突，请重试"]})
            continue

        created_at_value = _parse_dt(row.order_time) or beijing_now()
        order_days_value = calculate_order_days(created_at_value)
        gross_profit_value = await calc_order_gross_profit(row.product_name, row.sales_amount, db)

        new_order = Order(
            order_id=order_id,
            shop_id=shop.shop_id,
            product_name=row.product_name,
            platform_order_no=row.platform_order_no,
            sales_amount=row.sales_amount,
            freight=row.freight,
            shipping_status=row.shipping_status,
            produce_status=row.produce_status,
            produce_status_update_at=beijing_now() if row.produce_status != "unproduce" else None,
            produce_status_update_user=row.imported_by if row.produce_status != "unproduce" else None,
            logistics_company=row.logistics_company,
            logistics_no=row.logistics_no,
            receiver_address=row.receiver_address,
            remark=row.remark,
            refund_note=row.refund_note,
            commission_rate=None,  # 合并时按导入人实时提成率
            commission_amount=None,
            created_by=row.imported_by,
            created_at=created_at_value,
            order_days=order_days_value,
            gross_profit=gross_profit_value,
        )
        db.add(new_order)
        await db.flush()

        # 提成金额：导入人当前提成率
        owner = (await db.execute(select(User).where(User.username == row.imported_by))).scalar_one_or_none()
        if owner and owner.commission_rate:
            try:
                sales = float(row.sales_amount or 0)
                new_order.commission_rate = owner.commission_rate
                new_order.commission_amount = str(round(sales * owner.commission_rate / 100, 2))
            except (TypeError, ValueError):
                pass

        # 通知工厂端（与手工建单一致）
        creator_name = owner.real_name if owner else (row.imported_by or "")
        await NotificationService.send_order_created_notification(db, new_order, creator_name or row.imported_by)

        # 删除临时记录
        row.status = "merged"
        row.merged_order_id = order_id
        row.merged_at = beijing_now()
        await db.delete(row)
        merged_ok.append({"id": row.id, "platform_order_no": row.platform_order_no, "order_id": order_id})

    await db.commit()

    db.add(OperationLog(
        username=current_user.username,
        operation_type="订单导入合并",
        operation_content=f"审核合并 {len(merged_ok)} 条（失败 {len(merged_fail)} 条）：{', '.join(o['order_id'] for o in merged_ok[:5])}",
    ))
    await db.commit()

    return {
        "merged": len(merged_ok),
        "failed": len(merged_fail),
        "merged_items": merged_ok,
        "failed_items": merged_fail,
    }


@router.delete("/{row_id}")
async def delete_order_import(
    row_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """删除单条临时记录（异常单也可手动删除）。"""
    _check_import_perm(current_user)
    ensure_data_permission(current_user, '/order-imports', 'delete')
    row = (await db.execute(select(OrderImport).where(OrderImport.id == row_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="临时订单不存在")
    _check_row_visible(row, current_user)
    await db.delete(row)
    await db.commit()
    return {"message": "删除成功"}


@router.post("/delete-batch")
async def delete_batch_order_imports(
    payload: DeleteBatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """批量删除临时记录。"""
    _check_import_perm(current_user)
    ensure_data_permission(current_user, '/order-imports', 'delete')
    ids = payload.ids
    if not ids:
        raise HTTPException(status_code=400, detail="请先勾选要删除的记录")
    deleted = 0
    for rid in ids:
        row = (await db.execute(select(OrderImport).where(OrderImport.id == rid))).scalar_one_or_none()
        if not row:
            continue
        _check_row_visible(row, current_user)
        await db.delete(row)
        deleted += 1
    await db.commit()
    return {"message": f"已删除 {deleted} 条", "deleted": deleted}
